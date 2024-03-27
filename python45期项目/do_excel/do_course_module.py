#!/usr/bin/env python
# @Time     : 2024-03-16 14:46
# @Author   : 华
# @Email    : 867075698@qq.com
# @File     : do_course_module.py
# @Software : PyCharm

from openpyxl import load_workbook
from tools.do_ini import read_config
from tools.project_path import *


class Do_Course:
    @staticmethod
    def get_data(file_name):
        wb = load_workbook(file_name)
        mode = eval(read_config(case_config_path, "MODE", "mode"))
        test_data = []
        for key in mode:  # 遍历配置文件里面的字典--key是表单名(对每个表单进行遍历)
            sheet = wb[key]  # 表单名
            if mode[key] == 'all':
                for i in range(2, sheet.max_row + 1):
                    sub_data = {"case_id": sheet.cell(i, 1).value, "title": sheet.cell(i, 2).value,
                                "url": sheet.cell(i, 3).value, "data": sheet.cell(i, 4).value,
                                "method": sheet.cell(i, 5).value, "expected": sheet.cell(i, 6).value,
                                "result": sheet.cell(i, 7).value, "TestResult": sheet.cell(i, 8).value,
                                "sheet_name": key}  # "sheet_name": key  加一个表单名称
                    test_data.append(sub_data)
            else:
                for case_id in mode[key]:
                    sub_data = {"case_id": sheet.cell(case_id+1, 1).value, "title": sheet.cell(case_id+1, 2).value,
                                "url": sheet.cell(case_id+1, 3).value, "data": sheet.cell(case_id+1, 4).value,
                                "method": sheet.cell(case_id+1, 5).value, "expected": sheet.cell(case_id+1, 6).value,
                                "result": sheet.cell(case_id+1, 7).value, "TestResult": sheet.cell(case_id+1, 8).value}
                    test_data.append(sub_data)
        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)


if __name__ == '__main__':
    data = Do_Course.get_data(course_case_path)
    print(data)

